"""
functions for creating the inputsheets, and functions for calculating "umläufe"
functions for reading the TPT outputs
"""

import re
from typing import Tuple, Dict, Optional

import numpy as np
import pandas as pd
from ElevationSampler import ElevationProfile
from numpy import ndarray
from openpyxl import load_workbook
from pandas import DataFrame

from util import get_elevation_at_dist, get_maxspeed_at_dist, get_electrified_at_dist


def elevation_pipeline(elevation_profile: ElevationProfile, brunnels: DataFrame, first_sample_distance: float = 10.,
                       end_sample_distance: float = 100., resample: bool = True, resample_distance: float = 300.,
                       construct_brunnels: bool = True, max_brunnel_length: float = 300.,
                       construct_brunnel_thresh: float = 5., diff_kernel_dist: int = 10,
                       smooth_1: bool = True, smooth_window_size_1: int = 31, poly_order_1: int = 3,
                       smooth_2: bool = True, smooth_window_size_2: int = 11, poly_order_2: int = 1,
                       mode: str = "nearest", minimum: bool = True,
                       minimum_loops: int = 1, variance: bool = True, adjust_window_size: int = 12,
                       std_thresh: float = 2., sub_factor: float = 5., clip: float = 20, min_ele: float = -3,
                       max_ele: float = 2962.) -> ElevationProfile:

    # filter unrealistic values
    keep = (elevation_profile.elevations > min_ele) & (elevation_profile.elevations < max_ele)
    keep_orig = (elevation_profile.elevations_orig > min_ele) & (elevation_profile.elevations_orig < max_ele)

    elevation_profile.distances = elevation_profile.distances[keep]
    elevation_profile.elevations = elevation_profile.elevations[keep]

    elevation_profile.distances_orig = elevation_profile.distances_orig[keep_orig]
    elevation_profile.elevations_orig = elevation_profile.elevations_orig[keep_orig]

    if elevation_profile.distances.shape[0] <= 1:
        return elevation_profile

    # because of filter, brunnels can be in region where no elevation data
    # end_dist must be larger than last distance value
    brunnels = brunnels[brunnels.end_dist <= elevation_profile.distances[-1]]

    # first_resample so that equidistant sample points at first_sample_distance apart
    elevation_profile = elevation_profile.resample(first_sample_distance)

    # then interpolate the brunnels
    elevation_profile = elevation_profile.interpolate_brunnels(brunnels,
                                                               distance_delta=first_sample_distance,
                                                               construct_brunnels=construct_brunnels,
                                                               max_brunnel_length=max_brunnel_length,
                                                               construct_brunnel_thresh=construct_brunnel_thresh,
                                                               diff_kernel_dist=diff_kernel_dist)

    # then adjust the forest height
    if minimum:
        elevation_profile = elevation_profile.to_terrain_model(method="minimum", minimum_loops=minimum_loops)

    # then adjust the forest height again with method variance
    if variance:
        elevation_profile = elevation_profile.to_terrain_model(method="variance",
                                                               window_size=adjust_window_size,
                                                               std_thresh=std_thresh, sub_factor=sub_factor,
                                                               clip=clip)
    if resample:
        # then resample the elevation to resample_distance
        elevation_profile = elevation_profile.resample(resample_distance)

    if smooth_1:
        # then smooth the elevation profile with high polyorder
        elevation_profile = elevation_profile.smooth(window_size=smooth_window_size_1, poly_order=poly_order_1,
                                                     mode=mode)

    # then resample the elevation to the end sample distance
    elevation_profile = elevation_profile.resample(end_sample_distance)

    # then smooth again with averaging smoothing method
    if smooth_2:
        elevation_profile = elevation_profile.smooth(window_size=smooth_window_size_2, poly_order=poly_order_2,
                                                     mode=mode)

    return elevation_profile


def flip_trip(start_dists: ndarray, end_dists: ndarray, parameter: ndarray, invert_para: bool = False) \
        -> Tuple[ndarray, ndarray, ndarray]:
    """
    Calculate the distances and parameter from the other direction.

    Parameters
    ----------
    start_dists : Numpy Array
    end_dists : Numpy Array
    parameter : Numpy Array
    invert_para : bool
        Take parameter * -1

    Returns
    -------
    Numpy Array
        The flipped distances and parameter
    """

    trip_length = end_dists[-1]

    # calculate start dists from the end distances
    new_start_dists = end_dists[::-1]
    new_start_dists = trip_length - new_start_dists

    new_end_dists = start_dists[::-1]
    new_end_dists = trip_length - new_end_dists

    parameter = parameter[::-1]

    if invert_para:
        parameter = parameter * -1

    return new_start_dists, new_end_dists, parameter


def calc_end_dists(start_dists: ndarray, trip_length: float) -> ndarray:
    """

    Parameters
    ----------
    start_dists : 1d numpy array of floats
    trip_length : float

    Returns
    -------
        1d numpy array of floats
            The end distances
    """
    # end dist is always the start dist of next element. last element end dist is trip length
    end_dists = start_dists[1:].copy()
    end_dists = np.append(end_dists, trip_length)

    return end_dists


def create_umlauf(parameter_df: DataFrame, trip_count: int, parameter_column: str,
                  start_dists_column: str = "start_dist", end_dists_column: str = "end_dist", flip_first: bool = False,
                  drop_dups: bool = True):
    """
    For the given distances and parameter arrays returns the computed "umlauf" arrays.

    Parameters
    ----------

    parameter_df : pandas dataframe
        with parameter_column, start_dists_column and end_dists_column
    trip_count : int
    parameter_column : str
    start_dists_column : str
    end_dists_column : str
    flip_first : bool
    drop_dups : bool

    Returns
    -------
    pandas DataFrame
        The computed umlauf. Dataframe with start_dists_column and parameter columns.
    """

    invert_para = False

    if parameter_column == "incl":
        invert_para = True
        if end_dists_column not in parameter_df.columns:
            end_dists_column = start_dists_column

    start_dists = parameter_df[start_dists_column].values
    end_dists = parameter_df[end_dists_column].values
    parameter = parameter_df[parameter_column].values

    trip_length = end_dists[-1]

    if flip_first:
        start_dists, end_dists, parameter = flip_trip(start_dists, end_dists, parameter, invert_para=invert_para)

    res_dists = []
    res_para = []

    start_dist = 0
    for _ in range(trip_count):
        distances = start_dists + start_dist
        res_dists.append(distances)
        res_para.append(parameter)

        # add trip length
        start_dist += trip_length

        # flip the trip
        start_dists, end_dists, parameter = flip_trip(start_dists, end_dists, parameter, invert_para=invert_para)

    res_dists = np.concatenate(res_dists, axis=None)
    res_para = np.concatenate(res_para, axis=None)

    data = {start_dists_column: res_dists, parameter_column: res_para}
    res = pd.DataFrame(data)

    if drop_dups:
        res = res.drop_duplicates(ignore_index=True)

        drop_idx = []
        for i in range(1, res.shape[0]):
            if (res.iloc[i][parameter_column]) == (res.iloc[i - 1][parameter_column]):
                drop_idx.append(i)
        res = res.drop(drop_idx)
        res = res.reset_index(drop=True)

    return res


def trip_title_to_filename(title: str) -> str:
    special_char_map = {ord('ä'): 'ae', ord('ü'): 'ue', ord('ö'): 'oe', ord('ß'): 'ss', ord('('): '', ord(')'): '',
                        ord(' '): '_'}

    filename = title.lower()
    filename = filename.replace(" - ", " ")
    filename = filename.translate(special_char_map)
    filename = re.sub('[^a-z0-9_]+', '', filename)

    return filename


def write_sensor_input_sheet(trip_title: str, timetable: DataFrame, electrified: DataFrame, maxspeed: DataFrame,
                             inclination: DataFrame, params: Optional[Dict] = None,
                             folder: Optional[str] = None) -> str:
    """

    Parameters
    ----------
    trip_title
    timetable
    electrified
    maxspeed
    inclination
    params : dict
        sheet1 : "Timetable and limits",
        sheet2 : "Gradients and curves",
        sheet3 : "TPT requirements",
        accelerationLimits : 1.2
        decelerationLimits : 0.9
        gravity_acceleration : 9.81
        time_overhead : 0.05
        gradient_interpolation_binary : 0
        effort_in_curve : 8
        resistances_kp : "normal"
    folder : str
        the output folder

    Returns
    -------

    """

    default_params = {
        "sheet1": "Timetable and limits",
        "sheet2": "Gradients and curves",
        "sheet3": "TPT requirements",
        "accelerationLimits": 1.2,
        "decelerationLimits": 0.9,
        "gravity_acceleration": 9.81,
        "time_overhead": 0.05,
        "gradient_interpolation_binary": 0,
        "effort_in_curve": 8,
        "resistances_kp": "normal"
    }

    if params is not None:
        default_params.update(params)

    filename = trip_title_to_filename(trip_title)
    basename = filename
    if folder is not None:
        filename = folder + '/' + filename + '.xlsx'
    else:
        filename = filename + '.xlsx'

    writer = pd.ExcelWriter(filename, engine='xlsxwriter')

    # we first need to write a dataframe, to get the sheets later

    # timetable
    timetable.to_excel(writer, sheet_name=default_params['sheet1'], index=False, header=False, startrow=3)

    workbook = writer.book
    worksheet = writer.sheets[default_params['sheet1']]

    # timetable end
    current_row = timetable.shape[0] + 3
    worksheet.write(current_row, 0, "]end")
    current_row += 1

    # 2 empty rows
    current_row += 2

    # title
    worksheet.write('A1', ']title')
    worksheet.write('B1', trip_title)

    # timetable header
    worksheet.write(2, 0, "]timetable")
    worksheet.write(2, 1, "Station Name")
    worksheet.write(2, 2, "Stop time at station [s]")
    worksheet.write(2, 3, "Driving time to next station [s]")

    # electrification header
    worksheet.write(current_row, 0, "]electrification")
    worksheet.write(current_row, 1, "binary")
    current_row += 1

    # electrification
    electrified[["start_dist", "electrified"]].to_excel(writer, sheet_name=default_params['sheet1'], index=False,
                                                        header=False, startrow=current_row)

    current_row += electrified.shape[0]

    # electrification end
    worksheet.write(current_row, 0, "]end")
    current_row += 1

    # empty row
    current_row += 1

    # maxspeed header
    worksheet.write(current_row, 0, "]speedLimits")
    worksheet.write(current_row, 1, "v [km/h]")
    current_row += 1

    # maxspeed
    maxspeed[["start_dist", "maxspeed"]].to_excel(writer, sheet_name=default_params['sheet1'], index=False,
                                                  header=False, startrow=current_row)

    current_row += maxspeed.shape[0]

    # maxspeed end
    worksheet.write(current_row, 0, "]end")
    current_row += 1

    # 2 empty rows
    current_row += 2

    worksheet.write(current_row, 0, "]accelerationLimits")
    worksheet.write(current_row, 1, "a [m/s^2]")
    current_row += 1

    worksheet.write(current_row, 0, 0)
    worksheet.write(current_row, 1, default_params['accelerationLimits'])
    current_row += 1

    worksheet.write(current_row, 0, "]end")
    current_row += 1

    current_row += 1

    worksheet.write(current_row, 0, "]decelerationLimits")
    worksheet.write(current_row, 1, "a [m/s^2]")
    current_row += 1

    worksheet.write(current_row, 0, 0)
    worksheet.write(current_row, 1, default_params['decelerationLimits'])
    current_row += 1

    worksheet.write(current_row, 0, "]end")

    # column width
    worksheet.set_column(0, 3, 25)

    # new worksheet
    inclination.to_excel(writer, sheet_name=default_params['sheet2'], index=False, header=False, startrow=3)

    worksheet = writer.sheets[default_params['sheet2']]

    # title
    worksheet.write('A1', "]gravity_acceleration(m/s2)")
    worksheet.write('B1', default_params['gravity_acceleration'])

    # inclination header
    worksheet.write(2, 0, "]gradients")
    worksheet.write(2, 1, "Gradient [‰]")

    # inclination end
    current_row = inclination.shape[0] + 3
    worksheet.write(current_row, 0, "]end")
    current_row += 1

    # 2 empty rows
    current_row += 2

    worksheet.write(current_row, 0, "]curves")
    worksheet.write(current_row, 1, "Radius [m]")
    current_row += 1

    worksheet.write(current_row, 0, "]end")

    # column width
    worksheet.set_column(0, 3, 25)

    worksheet = workbook.add_worksheet(default_params['sheet3'])

    worksheet.write('A1', "]time_overhead")
    worksheet.write('B1', default_params['time_overhead'])

    worksheet.write('A4', "]gradient_interpolation_binary")
    worksheet.write('B4', default_params['gradient_interpolation_binary'])

    worksheet.write('A6', "]effort_in_curve(N.m/kg)")
    worksheet.write('B6', default_params['effort_in_curve'])

    worksheet.write('A9', "]resistances_kp(m)_name(word)")

    worksheet.write('A10', 0)
    worksheet.write('B10', default_params['resistances_kp'])

    worksheet.write('A11', "]end")

    worksheet.set_column(0, 3, 25)

    writer.save()

    return basename


def write_tpt_input_sheet(template_file: str, trip_title: str, timetable: DataFrame, electrified: DataFrame,
                          maxspeed: DataFrame, inclination: DataFrame, folder: Optional[str] = None) -> str:
    """

    Parameters
    ----------
    template_file : str
        the path to a tpt template .xlsx file
        in the template the rows must be cleared (see example template file in this module)
    trip_title
    timetable
    electrified
    maxspeed
    inclination
    folder : str
        the output folder

    Returns
    -------

    """

    timetable = timetable.copy()

    filename = trip_title_to_filename(trip_title)
    basename = filename
    if folder is not None:
        filename = folder + '/' + filename + '.xlsx'
    else:
        filename = filename + '.xlsx'

    wb = load_workbook(filename=template_file)

    # connect pandas writer with the workbook
    writer = pd.ExcelWriter(template_file, engine='openpyxl')
    writer.book = wb

    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

    ws = wb['driving']

    # timetable
    start_row, end_row = _get_table_row_boundaries(ws, "]stations_middle_kp(m)_(S/P)_t(s)_name(text)")
    start_row += 1

    # delete old timetable data
    ws.delete_rows(start_row, end_row - start_row)

    # write timetable
    # first insert the correct number of rows
    ws.insert_rows(start_row, timetable.shape[0])

    # make timetable the correct format
    timetable["s"] = "s"
    timetable = timetable[["dist", "s", "stop_duration", "stop_name", "driving_time"]]
    timetable.to_excel(writer, "driving", startrow=start_row - 1, header=False, index=False)

    # maxspeed
    start_row, end_row = _get_table_row_boundaries(ws, "]max_speeds_head_and_tail_kp(m)_v(km/h)")
    start_row += 1

    # delete old maxspeed data
    ws.delete_rows(start_row, end_row - start_row)

    # write maxspeed
    # first insert the correct number of rows
    ws.insert_rows(start_row, maxspeed.shape[0])

    # make maxspeed the correct format
    maxspeed = maxspeed[["start_dist", "maxspeed"]]
    maxspeed.to_excel(writer, "driving", startrow=start_row - 1, header=False, index=False)

    ws = wb['line']

    # title
    ws['B1'] = trip_title

    #
    #  electrified
    start_row, end_row = _get_table_row_boundaries(ws, "]electrification_kp(m)_binary_link(0,1)")
    start_row += 1

    # delete old electrified data
    ws.delete_rows(start_row, end_row - start_row)

    # write electrified
    # first insert the correct number of rows
    ws.insert_rows(start_row, electrified.shape[0])

    # make electrified the correct format
    electrified = electrified[["start_dist", "electrified"]]
    electrified.to_excel(writer, "line", startrow=start_row - 1, header=False, index=False)

    #
    #  inclination
    start_row, end_row = _get_table_row_boundaries(ws, "]gradients_kp(m)_g(o/oo)_link(0,1)")
    start_row += 1

    # delete old inclination data
    ws.delete_rows(start_row, end_row - start_row)

    # write inclination
    # first insert the correct number of rows
    ws.insert_rows(start_row, inclination.shape[0])

    # make inclination the correct format
    inclination["0"] = 0
    inclination = inclination[["start_dist", "incl", "0"]]
    inclination.to_excel(writer, "line", startrow=start_row - 1, header=False, index=False)

    wb.save(filename)

    return basename


def read_tpt_input_sheet(file) -> Tuple[DataFrame, DataFrame, DataFrame, DataFrame]:
    """
    Read TPT input file. Return electrified, maxspeed, timetable, elevation dataframes.

    Parameters
    ----------
    file

    Returns
    -------
    electrified, maxspeed, timetable, elevation dataframes

    """
    wb = load_workbook(filename=file)

    ws = wb['driving']

    # timetable
    start_row, end_row = _get_table_row_boundaries(ws, "]stations_middle_kp(m)_(S/P)_t(s)_name(text)")
    timetable = pd.read_excel(file, sheet_name='driving', skiprows=start_row, header=None,
                              names=['dist', 'stop_duration', 'stop_name', 'driving_time'],
                              nrows=end_row - start_row - 1, usecols='A,C:E')
    timetable['arrival_time'] = pd.to_timedelta((timetable.driving_time + timetable.stop_duration).cumsum(),
                                                unit='s').shift(1)
    timetable['departure_time'] = timetable.arrival_time + pd.to_timedelta(timetable.stop_duration, unit='s')
    timetable.at[0, 'departure_time'] = pd.to_timedelta(0, unit='s')
    timetable.at[0, 'arrival_time'] = pd.to_timedelta(0, unit='s')
    timetable['departure_time'] = pd.Timestamp(year=2021, month=1, day=1, hour=0, minute=0, second=0, microsecond=0,
                                               nanosecond=0) + timetable.departure_time
    timetable['arrival_time'] = pd.Timestamp(year=2021, month=1, day=1, hour=0, minute=0, second=0, microsecond=0,
                                             nanosecond=0) + timetable.arrival_time
    timetable = timetable[['dist', 'stop_name', 'stop_duration', 'driving_time', 'arrival_time', 'departure_time']]

    # maxspeed
    start_row, end_row = _get_table_row_boundaries(ws, "]max_speeds_head_and_tail_kp(m)_v(km/h)")
    maxspeed = pd.read_excel(file, sheet_name='driving', skiprows=start_row, header=None,
                             names=['start_dist', 'maxspeed'], nrows=end_row - start_row - 1, usecols='A,B')

    ws = wb['line']

    # Electrified
    start_row, end_row = _get_table_row_boundaries(ws, "]electrification_kp(m)_binary_link(0,1)")
    electrified = pd.read_excel(file, sheet_name='line', skiprows=start_row, header=None,
                                names=['start_dist', 'electrified'], nrows=end_row - start_row - 1, usecols='A,B')

    # Elevation
    start_row, end_row = _get_table_row_boundaries(ws, "]gradients_kp(m)_g(o/oo)_link(0,1)")
    elevation = pd.read_excel(file, sheet_name='line', skiprows=start_row, header=None, names=['distance', 'elevation'],
                              nrows=end_row - start_row - 1, usecols='A,B')
    elevation['elevation'] = (elevation.elevation / 10).cumsum()
    elevation['elevation'] += elevation.elevation.abs().max()
    trip_length = elevation.distance.iloc[-1]

    maxspeed['end_dist'] = maxspeed.start_dist.shift(-1)
    maxspeed.iloc[-1, 2] = trip_length
    maxspeed = maxspeed[['maxspeed', 'start_dist', 'end_dist']]

    electrified['end_dist'] = electrified.start_dist.shift(-1)
    electrified.iloc[-1, 2] = trip_length
    electrified = electrified[['electrified', 'start_dist', 'end_dist']]

    return electrified, maxspeed, timetable, elevation


def _get_table_row_boundaries(ws, begin_token: str, end_token: str = "]end"):
    start_row = None
    end_row = None

    # get the rows where table data is saved
    in_section = False
    for row in ws.iter_rows(min_col=0, max_col=0):
        cell = row[0]
        if cell.value == begin_token:
            start_row = cell.row
            in_section = True
        if in_section and cell.value == end_token:
            end_row = cell.row
            break

    return start_row, end_row


def read_tpt_output_sheet(file, version=2):

    if version == 1:
        tpt_df = pd.read_csv(file, header=None,
                             names=['time', 'acceleration', 'velocity', 'distance', 'force', 'power'])
        tpt_df = tpt_df[['time', 'distance', 'acceleration', 'velocity', 'force', 'power']]
    else:
        tpt_df = pd.read_csv(file, header=0,
                             names=['time', 'acceleration', 'velocity', 'distance', 'force', 'power', 'maxspeed_sim',
                                    'gradient', 'electrification'])
        tpt_df = tpt_df[['time', 'distance', 'acceleration', 'velocity', 'force', 'power', 'maxspeed_sim', 'gradient',
                         'electrification']]

    # convert to km/h
    tpt_df['velocity'] = tpt_df['velocity'] * 3.6

    if 'maxspeed_sim' in tpt_df.columns:
        tpt_df['maxspeed_sim'] = tpt_df['maxspeed_sim'] * 3.6

    # add time data

    # convert index to timedelta64
    # add time from timetable, possibly add second datetime64 index / or as column?

    tpt_df["time_delta"] = pd.to_timedelta(tpt_df["time"], unit='S')

    tpt_df.set_index("time_delta", inplace=True, drop=False)

    return tpt_df


def add_inputs_to_simulation_results(tpt_df, elevation, maxspeed, electrified):
    # resample the inputs so the distances match the simulation distances

    tpt_df[['elevation', 'maxspeed', 'electrified']] = tpt_df.apply(
        lambda r: pd.Series([get_elevation_at_dist(elevation, r["distance"]),
                             get_maxspeed_at_dist(maxspeed, r["distance"]),
                             get_electrified_at_dist(electrified, r["distance"])],
                            index=['elevation', 'maxspeed', 'electrified']), axis=1)
    return tpt_df


def resample_simulation_results(tpt_df, t: Optional[int] = 10):
    # mean: distance, acceleration, velocity, elevation
    # median: maxspeed, electrified
    # sum: force, power
    # min: time

    # skip if already resampled to given interval
    if tpt_df.iloc[1]["time_delta"] == pd.Timedelta(t, unit="S"):
        return tpt_df

    interval = str(t) + "S"

    resample = tpt_df.resample(interval)
    resampled = resample[["time_delta"]].min()

    if "time" in tpt_df.columns:
        resampled[["time"]] = resample[["time"]].min()

    if "distance" in tpt_df.columns:
        resampled[["distance"]] = resample[["distance"]].mean()

    if "acceleration" in tpt_df.columns:
        resampled[["acceleration"]] = resample[["acceleration"]].mean()

    if "velocity" in tpt_df.columns:
        resampled[["velocity"]] = resample[["velocity"]].mean()

    if "power" in tpt_df.columns:
        resampled[["power"]] = resample[["power"]].mean()

    if "force" in tpt_df.columns:
        resampled[["force"]] = resample[["force"]].mean()

    if "elevation" in tpt_df.columns:
        resampled[["elevation"]] = resample[["elevation"]].mean()

    if "gradient" in tpt_df.columns:
        resampled[["gradient"]] = resample[["gradient"]].mean()

    if "maxspeed" in tpt_df.columns:
        resampled[["maxspeed"]] = resample[["maxspeed"]].median()

    if "maxspeed_sim" in tpt_df.columns:
        resampled[["maxspeed_sim"]] = resample[["maxspeed_sim"]].median()

    if "electrified" in tpt_df.columns:
        resampled[["electrified"]] = resample[["electrified"]].median()

    if "electrification" in tpt_df.columns:
        resampled[["electrification"]] = resample[["electrification"]].median()

    return resampled

